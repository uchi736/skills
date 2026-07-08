# -*- coding: utf-8 -*-
"""
corp_xlsx — 自社ハウススタイルの Excel 生成ライブラリ(corp-excel スキル同梱)

使い方(最小):
    import sys; sys.path.insert(0, os.path.expanduser("~/.claude/skills/corp-excel/scripts"))
    from corp_xlsx import *
    wb = new_book()
    ws = wb.active; ws.title = "サマリ"
    setup_sheet(ws)
    sheet_title(ws, "FY2026 業績サマリ", span=8)
    kpi_row(ws, "B4", [("売上収益", 1500, NUM, "+5.6%"), ("営業利益", 120, NUM, "+12.0%")])
    write_table(ws, "B9", ["セグメント", "FY25", "FY26計画"], rows, formats=[None, NUM, NUM])
    wb.save("out.xlsx")

生成物は Excel 上で完全編集可能。検証は export_preview.py(Excel COM → PDF → PNG)で行う。
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.properties import PageSetupProperties

# ============================================================ パレット(corp-slides と共通)
NAVY   = "01357A"
BLUE   = "004EA2"
BLUE_TX= "00549A"
STEEL  = "3377AE"
CYAN   = "00BCEB"
CYANSUB= "66C5DC"
TEAL   = "00BB99"
RED    = "C00000"
BG     = "F0F5F8"
LTBLUE = "D6ECF7"
INK    = "262626"
INK2   = "4D4D4D"
GREY   = "A6A6A6"
GRID   = "E3E7EC"
FONT   = "Yu Gothic"

# ============================================================ 数値書式(この定数を使う)
NUM   = '#,##0;[Red]"△"#,##0'          # 整数・負値は赤△(日本の財務標準)
NUM1  = '#,##0.0;[Red]"△"#,##0.0'      # 小数1桁
PCT   = '0.0%;[Red]"△"0.0%'            # パーセント(値は 0.056 のような比率で渡す)
PCT_PT= '+0.0"pt";[Red]"△"0.0"pt"'     # ポイント差
YEN_MM= '#,##0,,"百万円";[Red]"△"#,##0,,"百万円"'  # 円単位の値を百万円表示(表示のみスケール)
DATE  = 'yyyy/m/d'
PLAIN = '#,##0'                          # 負値も黒のまま(チャート軸など)

_thin  = Side(style="thin",  color=GREY)
_hair  = Side(style="hair",  color=GREY)
_med   = Side(style="medium", color=NAVY)
_dbl   = Side(style="double", color=NAVY)

# ブランド固有値(社名フッター等)は assets/brand/brand.json(git管理外)から。無ければ中立既定。
import os as _os


def _load_brand():
    import json
    try:
        with open(_os.path.join(_os.path.dirname(_os.path.abspath(__file__)),
                                "..", "assets", "brand", "brand.json"), encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

_BRAND = _load_brand()
FOOTER = _BRAND.get("footer", "© Your Company All Rights Reserved.")


def _f(size=11, bold=False, color=INK, name=FONT):
    return Font(name=name, size=size, bold=bold, color=color)


def _fill(hex6):
    return PatternFill("solid", start_color=hex6, end_color=hex6)


def new_book():
    """既定フォントを游ゴシックにしたブックを作る"""
    wb = Workbook()
    for st in wb._named_styles:
        if st.name == "Normal":
            st.font = _f()
    return wb


def setup_sheet(ws, tab_color=BLUE, gridlines=False, landscape=True):
    """提出用シートの共通設定: 枠線非表示・タブ色・A4横・幅1ページ印刷・©フッター"""
    ws.sheet_view.showGridLines = gridlines
    ws.sheet_properties.tabColor = tab_color
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.oddFooter.left.text = FOOTER
    ws.oddFooter.left.size = 8
    ws.oddFooter.right.text = "&P / &N"
    ws.oddFooter.right.size = 8
    ws.column_dimensions["A"].width = 2.5  # 左余白列(A列は使わない)


def sheet_title(ws, title, cell="B2", sub=None, span=8):
    """シート見出し: 太字ブルー + 濃紺アンダーライン(スライドの見出しと同じ意匠)"""
    col, row = coordinate_from_string(cell)
    c0 = column_index_from_string(col)
    tc = ws.cell(row=row, column=c0, value=title)
    tc.font = _f(14, True, BLUE_TX)
    ws.row_dimensions[row].height = 22
    for i in range(span):
        ws.cell(row=row, column=c0 + i).border = Border(bottom=_med)
    if sub:
        sc = ws.cell(row=row + 1, column=c0, value=sub)
        sc.font = _f(9, False, INK2)
    return row + (2 if sub else 1)


def unit_note(ws, cell, text="(単位:百万円)"):
    col, row = coordinate_from_string(cell)
    c = ws.cell(row=row, column=column_index_from_string(col), value=text)
    c.font = _f(9, False, INK2)
    c.alignment = Alignment(horizontal="right")


def write_table(ws, anchor, headers, rows, formats=None, widths=None,
                zebra=True, freeze=True, total_row=False, first_col_header=True,
                header_height=20, autofilter=False):
    """
    自社スタイルの表を書く。
    anchor : 左上セル(例 "B5")
    formats: 列ごとの数値書式(None=文字列列)。例 [None, NUM, NUM, PCT]
    widths : 列幅(文字数)。例 [18, 12, 12, 10]
    total_row=True で最終行を合計行として装飾(二重罫線+太字)
    戻り値: dict(header_row, first_row, last_row, first_col, last_col) — チャート参照用
    """
    col, row = coordinate_from_string(anchor)
    c0 = column_index_from_string(col)
    ncols = len(headers)
    formats = formats or [None] * ncols
    # ヘッダ
    ws.row_dimensions[row].height = header_height
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=c0 + j, value=h)
        c.font = _f(11, True, "FFFFFF")
        c.fill = _fill(BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=_med)
    # 本体
    n = len(rows)
    for i, r in enumerate(rows):
        rr = row + 1 + i
        is_total = total_row and i == n - 1
        for j, v in enumerate(r):
            c = ws.cell(row=rr, column=c0 + j, value=v)
            fmt = formats[j] if j < len(formats) else None
            bold = is_total or (j == 0 and first_col_header)
            c.font = _f(11, bold, INK)
            if fmt:
                c.number_format = fmt
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if is_total:
                c.fill = _fill(LTBLUE)
                c.border = Border(top=_dbl, bottom=_thin)
            else:
                if j == 0 and first_col_header:
                    c.fill = _fill(LTBLUE if (zebra and i % 2 == 1) else "E8F2FA")
                elif zebra and i % 2 == 1:
                    c.fill = _fill(BG)
                c.border = Border(bottom=_hair)
    # 列幅
    if widths:
        for j, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(c0 + j)].width = w
    if freeze:
        ws.freeze_panes = ws.cell(row=row + 1, column=c0).coordinate
    if autofilter:
        last = ws.cell(row=row + n, column=c0 + ncols - 1).coordinate
        ws.auto_filter.ref = f"{anchor}:{last}"
    return dict(header_row=row, first_row=row + 1, last_row=row + n,
                first_col=c0, last_col=c0 + ncols - 1)


def kpi_row(ws, anchor, kpis, block_w=3, gap=1):
    """
    KPIタイル列: kpis = [(ラベル, 値, 数値書式, 前年比等の補足|None), ...]
    補足は "+5.6%" → 緑 / "△1.2%" "-3%" → 赤 で自動色分け
    """
    col, row = coordinate_from_string(anchor)
    c0 = column_index_from_string(col)
    for k, item in enumerate(kpis):
        label, value, fmt, delta = (list(item) + [None] * 4)[:4]
        left = c0 + k * (block_w + gap)
        right = left + block_w - 1
        rng = lambda r: f"{get_column_letter(left)}{r}:{get_column_letter(right)}{r}"
        # ラベル(上端に青のアクセント罫線)
        ws.merge_cells(rng(row))
        lc = ws.cell(row=row, column=left, value=label)
        lc.font = _f(10, True, INK)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        # 値
        ws.merge_cells(rng(row + 1))
        vc = ws.cell(row=row + 1, column=left, value=value)
        vc.font = _f(20, True, BLUE)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            vc.number_format = fmt
        ws.row_dimensions[row + 1].height = 30
        # 補足
        ws.merge_cells(rng(row + 2))
        if delta is not None:
            color = INK2
            d = str(delta)
            if d.startswith(("+", "▲")):
                color = TEAL
            elif d.startswith(("-", "−", "△", "▼")):
                color = RED
            dc = ws.cell(row=row + 2, column=left, value=d)
            dc.font = _f(10, True, color)
            dc.alignment = Alignment(horizontal="center", vertical="center")
        # タイル装飾(背景+上端アクセント)
        for rr in range(row, row + 3):
            for cc in range(left, right + 1):
                cell = ws.cell(row=rr, column=cc)
                if not cell.fill or cell.fill.start_color.rgb in (None, "00000000"):
                    cell.fill = _fill(BG)
                top = Side(style="thick", color=BLUE) if rr == row else None
                cell.border = Border(top=top)
    return row + 3


def add_chart(ws, kind, anchor, data_ws, table_info, title=None,
              value_fmt=PLAIN, width=16, height=8.5, series_colors=None):
    """
    write_table の戻り値(table_info)からチャートを作って ws の anchor に置く。
    kind: "bar" | "line" | "stack"
    1列目=カテゴリ、2列目以降=系列 として扱う。
    """
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.chart.axis import ChartLines
    from openpyxl.drawing.line import LineProperties
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import (RichTextProperties, Paragraph,
                                       ParagraphProperties, CharacterProperties,
                                       Font as XFont)
    colors = series_colors or [BLUE, CYANSUB, TEAL, STEEL, NAVY]
    ti = table_info
    if kind == "line":
        ch = LineChart()
    else:
        ch = BarChart()
        ch.type = "col"
        ch.grouping = "stacked" if kind == "stack" else "clustered"
        if kind == "stack":
            ch.overlap = 100
        ch.gapWidth = 60
    data = Reference(data_ws, min_col=ti["first_col"] + 1, max_col=ti["last_col"],
                     min_row=ti["header_row"], max_row=ti["last_row"])
    cats = Reference(data_ws, min_col=ti["first_col"], max_col=ti["first_col"],
                     min_row=ti["first_row"], max_row=ti["last_row"])
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    # スタイル
    def _txpr(sz=900):
        cp = CharacterProperties(sz=sz, latin=XFont(typeface=FONT), ea=XFont(typeface=FONT))
        return RichText(bodyPr=RichTextProperties(),
                        p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    for i, s in enumerate(ch.series):
        gp = GraphicalProperties(solidFill=colors[i % len(colors)])
        if kind == "line":
            gp = GraphicalProperties(ln=LineProperties(solidFill=colors[i % len(colors)], w=28575))
        s.graphicalProperties = gp
        s.smooth = False
    ch.y_axis.majorGridlines = ChartLines(
        spPr=GraphicalProperties(ln=LineProperties(solidFill=GRID, w=9525)))
    ch.y_axis.numFmt = value_fmt
    if kind != "line":
        ch.y_axis.scaling.min = 0  # 棒グラフの0始まりは省略不可(差の誇張防止)
    ch.y_axis.delete = False
    ch.x_axis.delete = False
    ch.y_axis.txPr = _txpr()
    ch.x_axis.txPr = _txpr()
    ch.y_axis.spPr = GraphicalProperties(ln=LineProperties(noFill=True))
    ch.x_axis.spPr = GraphicalProperties(ln=LineProperties(solidFill=GREY, w=9525))
    if title:
        ch.title = title
    if len(ch.series) <= 1:
        ch.legend = None
    else:
        ch.legend.position = "b"
        ch.legend.overlay = False  # 軸ラベルとの重なり防止
    ch.width = width
    ch.height = height
    ws.add_chart(ch, anchor)
    return ch
